# import openpyxl to read/write Excel files
import openpyxl

# used to color cells (green / red)
from openpyxl.styles import PatternFill


# -------------------------------------------------
# get total number of rows in sheet
# used for loop in data driven testing
# -------------------------------------------------
def get_row_count(file, sheetname):

    # open excel file
    workbook = openpyxl.load_workbook(file)

    # select sheet
    sheet = workbook[sheetname]

    # return total rows
    return sheet.max_row



# -------------------------------------------------
# get total number of columns
# not always needed but useful
# -------------------------------------------------
def get_column_count(file, sheetname):

    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]

    return sheet.max_column



# -------------------------------------------------
# read data from excel
# used to get username / password / expected result
# -------------------------------------------------
def read_data(file, sheetname, row_num, column_num):

    workbook = openpyxl.load_workbook(file)

    sheet = workbook[sheetname]

    # return value from cell
    return sheet.cell(row=row_num, column=column_num).value



# -------------------------------------------------
# write data into excel
# used to write PASS / FAIL
# -------------------------------------------------
def write_data(file, sheetname, row_num, column_num, data):

    workbook = openpyxl.load_workbook(file)

    sheet = workbook[sheetname]

    # write data in cell
    sheet.cell(row=row_num, column=column_num).value = data

    # save file
    workbook.save(file)



# -------------------------------------------------
# fill cell with green color (PASS)
# -------------------------------------------------
def fill_green(file, sheetname, row_num, column_num):

    workbook = openpyxl.load_workbook(file)

    sheet = workbook[sheetname]

    # green color style
    greenFill = PatternFill(
        start_color='00FF00',
        end_color='00FF00',
        fill_type='solid'
    )

    # apply color to cell
    sheet.cell(row=row_num, column=column_num).fill = greenFill

    workbook.save(file)



# -------------------------------------------------
# fill cell with red color (FAIL)
# -------------------------------------------------
def fill_red(file, sheetname, row_num, column_num):

    workbook = openpyxl.load_workbook(file)

    sheet = workbook[sheetname]

    # red color style
    redFill = PatternFill(
        start_color='FF0000',
        end_color='FF0000',
        fill_type='solid'
    )

    # apply color
    sheet.cell(row=row_num, column=column_num).fill = redFill

    workbook.save(file)